"""
Report export service. Generates Excel (primary, via openpyxl) and PDF
(secondary, via reportlab) files from a list of parking sessions.
"""
import io
from datetime import date
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from app.models.parking_session import ParkingSession
from app.utils.time import duration_hours, format_duration, to_display_tz

HEADERS = [
    "Truck Number", "Driver Mobile", "Driver Name", "Transport Company",
    "Entry Time", "Exit Time", "Duration", "Status", "Amount", "Payment Mode", "Payment Status",
]


def _row_for_session(session: ParkingSession) -> list:
    truck = session.truck
    payment = session.payment
    dur = (
        format_duration(duration_hours(session.entry_time, session.exit_time))
        if session.exit_time else "—"
    )
    return [
        truck.truck_number,
        truck.driver_mobile,
        truck.driver_name or "",
        truck.transport_company or "",
        to_display_tz(session.entry_time).strftime("%d-%b-%Y %I:%M %p"),
        to_display_tz(session.exit_time).strftime("%d-%b-%Y %I:%M %p") if session.exit_time else "",
        dur,
        session.status.value,
        float(payment.amount) if payment else "",
        payment.payment_mode.value if payment and payment.payment_mode else "",
        payment.payment_status.value if payment else "",
    ]


def export_sessions_to_excel(
    sessions: Sequence[ParkingSession], from_date: date, to_date: date
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sessions"

    ws.append([f"Truck Parking Report: {from_date} to {to_date}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(HEADERS)

    header_row = ws.max_row
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    total_revenue = 0.0
    for session in sessions:
        ws.append(_row_for_session(session))
        if session.payment and session.payment.payment_status.value == "paid":
            total_revenue += float(session.payment.amount)

    ws.append([])
    ws.append(["Total Sessions", len(sessions)])
    ws.append(["Total Revenue (Paid)", total_revenue])

    for col_idx, header in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(header) + 4)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def export_sessions_to_pdf(
    sessions: Sequence[ParkingSession], from_date: date, to_date: date
) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=10 * mm, rightMargin=10 * mm, topMargin=10 * mm, bottomMargin=10 * mm,
    )
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(f"Truck Parking Report: {from_date} to {to_date}", styles["Title"]),
        Spacer(1, 6),
    ]

    data = [HEADERS] + [_row_for_session(s) for s in sessions]
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a2e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f2f2")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ])
    )
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()
